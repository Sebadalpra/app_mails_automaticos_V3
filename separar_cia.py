import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import customtkinter as ctk

def separar_cias(nueva_ventana, archivo_separacion):
    # Lee los datos desde el archivo proporcionado
    if archivo_separacion:
        archivo_excel = pd.read_excel(archivo_separacion)
        ctk.CTkLabel(nueva_ventana, text="Archivo Excel cargado", fg_color="green").pack(pady=5)
    else:
        archivo_excel = None
        ctk.CTkLabel(nueva_ventana, text="No se seleccionó ningún archivo", fg_color="red").pack(pady=5)
        return

    carpeta_salida = 'SIEP'
    if not os.path.exists(carpeta_salida):
        os.makedirs(carpeta_salida)

    # Obtén una lista única de los códigos de CIA
    codigos_cia_unicos = archivo_excel['CIA'].unique()

    # Itera a través de los códigos de CIA y guarda los subconjuntos en archivos separados
    for codigo_cia in codigos_cia_unicos:
        # Filtra por el código de CIA y trae todas las columnas disponibles excepto 'CORREOS' y 'NOMBRE_ARCHIVO'
        subset = archivo_excel[archivo_excel['CIA'] == codigo_cia]
        
        # Excluir las columnas 'CORREOS' y 'NOMBRE_ARCHIVO' si están presentes
        if 'CORREOS' in subset.columns:
            subset = subset.drop(columns=['CORREOS'])
        if 'NOMBRE_ARCHIVO' in subset.columns:
            subset = subset.drop(columns=['NOMBRE_ARCHIVO'])
        
        # Asegura que el código de CIA tenga un formato de 4 caracteres con ceros a la izquierda
        codigo_cia_formateado = f'{codigo_cia:04}'
        nombre_archivo = os.path.join(carpeta_salida, f'{codigo_cia_formateado}.xlsx')
        
        # Crea un nuevo libro de trabajo de Excel
        wb = Workbook()
        ws = wb.active
        
        # Configura los encabezados de columna
        for col_idx, col_name in enumerate(subset.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")  # Hace que el encabezado sea en negrita y color blanco
            cell.alignment = Alignment(horizontal='center', vertical='center')  # Centra el texto
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Color de fondo azul
            
        # Convierte el subconjunto de pandas al formato de hoja de cálculo de Excel
        for r_idx, row in enumerate(subset.iterrows(), start=2):
            for c_idx, value in enumerate(row[1], start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center')  # Justifica a la izquierda
                if isinstance(value, (int, float)):  # Ajuste de formato si es número
                    cell.number_format = '#,##0'
        
        # Ajusta el ancho de las columnas para que el texto se ajuste completamente
        for col in ws.columns:
            max_length = max(len(str(cell.value)) for cell in col) + 2
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length

        wb.save(nombre_archivo)  # Guarda el archivo Excel

    print("CIAS SEPARADAS CORRECTAMENTE")